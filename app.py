from flask import Flask, render_template, request, redirect, send_file, session, abort, flash, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from urllib.parse import urlencode
from datetime import datetime
from functools import wraps
import sqlite3
import io
import json
import re
import secrets

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer


# =========================================================
# APP / CONFIG
# =========================================================
app = Flask(__name__)
DB_FILE = "combustivel.db"
app.secret_key = "Rs231120"

# CONFIG DE ASSINATURA
RESPONSAVEL_NOME = "ALVINA TUPINAMBA"
RESPONSAVEL_CARGO = "SECRETÁRIA"
RESPONSAVEL_SETOR = "SECRETARIA MUNICIPAL DE SAUDE"


# =========================================================
# CSRF (manual, leve) — GLOBAL (before_request)
# =========================================================
def _ensure_csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_urlsafe(32)
    return session["csrf_token"]


@app.context_processor
def inject_csrf():
    def csrf_token():
        return _ensure_csrf_token()

    def csrf_input():
        return f'<input type="hidden" name="csrf_token" value="{_ensure_csrf_token()}">'

    return dict(csrf_token=csrf_token, csrf_input=csrf_input)


@app.before_request
def csrf_protect():
    # garante que exista um token na sessão (evita 400 bobo)
    _ensure_csrf_token()

    # protege somente ações que alteram dados
    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        token_form = request.form.get("csrf_token")
        token_header = request.headers.get("X-CSRF-Token")
        token = token_form or token_header

        if not token or token != session.get("csrf_token"):
            abort(400)


# =========================================================
# FILTERS / HELPERS
# =========================================================
@app.template_filter("data_br")
def data_br(data_iso):
    if not data_iso:
        return ""
    try:
        ano, mes, dia = data_iso.split("-")
        return f"{dia}/{mes}/{ano}"
    except Exception:
        return data_iso


def normalizar(texto: str) -> str:
    return (texto or "").strip().upper()


def normalizar_mes_param(mes_raw: str) -> str:
    """
    Aceita:
      - 'YYYY-MM' (ex: 2026-01)
      - 'MM/YYYY' (ex: 01/2026)
    Retorna sempre 'YYYY-MM' ou ''.
    """
    if not mes_raw:
        return ""
    mes_raw = (mes_raw or "").strip()

    if re.match(r"^\d{4}-\d{2}$", mes_raw):
        return mes_raw

    m = re.match(r"^(\d{2})/(\d{4})$", mes_raw)
    if m:
        mm, yyyy = m.group(1), m.group(2)
        return f"{yyyy}-{mm}"

    return mes_raw


def mes_pt_ano(mes_yyyy_mm: str) -> str:
    meses = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    try:
        ano, mes = mes_yyyy_mm.split("-")
        mes_nome = meses[int(mes) - 1]
        return f"{mes_nome} / {ano}"
    except Exception:
        return mes_yyyy_mm


def agora_emissao_str():
    return datetime.now().strftime("%d/%m/%Y %H:%M")


def agora_iso():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# =========================================================
# POSTOS / TOTAIS
# =========================================================

POSTOS_VALIDOS = {"PONTAO", "ESTRADA", "CIDADE"}


def posto_label(p: str) -> str:
    p = (p or "CIDADE").strip().upper()

    if p == "PONTÃO":
        p = "PONTAO"

    if p == "PONTAO":
        return "Pontão"
    if p == "ESTRADA":
        return "Estrada"
    return "Cidade"


def buscar_totais_postos(mes: str, nome_f: str, categoria_f: str, equipe_f: str):
    """
    Retorna totais organizados por posto e total geral.
    Usado em PDF e Excel.
    """

    mes = (mes or "").strip()
    nome_f = (nome_f or "").strip()
    categoria_f = (categoria_f or "").strip()
    equipe_f = (equipe_f or "").strip()

    where = []
    params = []

    if mes:
        where.append("strftime('%Y-%m', e.data) = ?")
        params.append(mes)

    if nome_f:
        where.append("UPPER(c.nome) LIKE UPPER(?)")
        params.append(f"%{nome_f}%")

    if categoria_f:
        where.append("c.categoria = ?")
        params.append(categoria_f)

    if equipe_f:
        where.append("c.equipe = ?")
        params.append(equipe_f)

    where_sql = (" AND " + " AND ".join(where)) if where else ""

    sql = f"""
        SELECT
            IFNULL(e.posto,'CIDADE') as posto,
            e.combustivel,
            IFNULL(SUM(e.quantidade),0)
        FROM entregas e
        JOIN cadastrados c ON c.id = e.cadastrado_id
        WHERE 1=1
        {where_sql}
        GROUP BY IFNULL(e.posto,'CIDADE'), e.combustivel
    """

    base = {
        "CIDADE": {"gasolina": 0.0, "diesel": 0.0, "total": 0.0},
        "ESTRADA": {"gasolina": 0.0, "diesel": 0.0, "total": 0.0},
        "PONTAO": {"gasolina": 0.0, "diesel": 0.0, "total": 0.0},
        "_totais": {"gasolina": 0.0, "diesel": 0.0, "total": 0.0},
    }

    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()

    for posto, combustivel, total in rows:
        p = (posto or "CIDADE").strip().upper()
        if p == "PONTÃO":
            p = "PONTAO"
        if p not in POSTOS_VALIDOS:
            p = "CIDADE"

        c = (combustivel or "").lower()
        if c not in ("gasolina", "diesel"):
            continue

        v = float(total or 0)

        base[p][c] += v
        base[p]["total"] += v

        base["_totais"][c] += v
        base["_totais"]["total"] += v

    return base

# =========================================================
# AUDITORIA
# =========================================================
def usuario_atual():
    return session.get("user") or {}


def log_acao(acao: str, entidade: str = None, entidade_id: int = None, detalhes: dict = None):
    """Registra ação na auditoria (não quebra o sistema se falhar)."""
    try:
        u = usuario_atual()
        ip = request.headers.get("X-Forwarded-For", request.remote_addr)

        payload = None
        if detalhes is not None:
            try:
                payload = json.dumps(detalhes, ensure_ascii=False)
            except Exception:
                payload = str(detalhes)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO auditoria (datahora, user_id, username, nome_exibicao, role, acao, entidade, entidade_id, detalhes, ip)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            agora_iso(),
            u.get("id"),
            u.get("username"),
            u.get("nome_exibicao"),
            u.get("role"),
            acao,
            entidade,
            entidade_id,
            payload,
            ip
        ))
        conn.commit()
        conn.close()
    except Exception:
        pass


# =========================================================
# BANCO / MIGRAÇÕES
# =========================================================
def criar_banco():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS cadastrados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            categoria TEXT,
            equipe TEXT,
            cota REAL
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS entregas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cadastrado_id INTEGER,
            quantidade REAL,
            data TEXT,
            observacao TEXT,
            combustivel TEXT,
            posto TEXT NOT NULL DEFAULT 'CIDADE',
            FOREIGN KEY(cadastrado_id) REFERENCES cadastrados(id)
        )
        """)

        # Migração: entregas (adicionar coluna posto se o banco já existir antigo)
        cursor.execute("PRAGMA table_info(entregas)")
        cols_entregas = [r[1] for r in cursor.fetchall()]
        if "posto" not in cols_entregas:
            cursor.execute("ALTER TABLE entregas ADD COLUMN posto TEXT NOT NULL DEFAULT 'CIDADE'")

        # Normaliza valores antigos / inválidos
        cursor.execute("""
            UPDATE entregas
               SET posto = 'CIDADE'
             WHERE posto IS NULL OR TRIM(posto) = ''
        """)
        cursor.execute("""
            UPDATE entregas
               SET posto = 'PONTAO'
             WHERE UPPER(TRIM(posto)) IN ('PONTÃO', 'PONTAO')
        """)
        cursor.execute("""
            UPDATE entregas
               SET posto = UPPER(TRIM(posto))
             WHERE posto IS NOT NULL AND TRIM(posto) <> ''
        """)
        cursor.execute("""
            UPDATE entregas
               SET posto = 'CIDADE'
             WHERE posto NOT IN ('PONTAO', 'ESTRADA', 'CIDADE')
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS estoque_mensal (
            mes TEXT NOT NULL,
            combustivel TEXT NOT NULL,
            limite REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (mes, combustivel)
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS categorias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS equipes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        )
        """)

        # USUÁRIOS
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            senha TEXT NOT NULL,
            role TEXT NOT NULL,
            nome_exibicao TEXT,
            ativo INTEGER NOT NULL DEFAULT 1,
            criado_em TEXT
        )
        """)

        # Migrações usuarios
        cursor.execute("PRAGMA table_info(usuarios)")
        cols = [r[1] for r in cursor.fetchall()]
        if "nome_exibicao" not in cols:
            cursor.execute("ALTER TABLE usuarios ADD COLUMN nome_exibicao TEXT")
        if "ativo" not in cols:
            cursor.execute("ALTER TABLE usuarios ADD COLUMN ativo INTEGER NOT NULL DEFAULT 1")
        if "criado_em" not in cols:
            cursor.execute("ALTER TABLE usuarios ADD COLUMN criado_em TEXT")

        cursor.execute("""
            UPDATE usuarios
               SET nome_exibicao = username
             WHERE nome_exibicao IS NULL OR nome_exibicao = ''
        """)
        cursor.execute("""
            UPDATE usuarios
               SET criado_em = ?
             WHERE criado_em IS NULL OR criado_em = ''
        """, (agora_iso(),))

        # AUDITORIA
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS auditoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datahora TEXT NOT NULL,
            user_id INTEGER,
            username TEXT,
            nome_exibicao TEXT,
            role TEXT,
            acao TEXT NOT NULL,
            entidade TEXT,
            entidade_id INTEGER,
            detalhes TEXT,
            ip TEXT
        )
        """)

        # AUDITORIA ADMIN (registro de limpezas)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS auditoria_admin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datahora TEXT NOT NULL,
            user_id INTEGER,
            username TEXT,
            nome_exibicao TEXT,
            role TEXT,
            acao TEXT NOT NULL,
            detalhes TEXT,
            ip TEXT
        )
        """)

        # ÍNDICES (performance)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_entregas_data ON entregas(data)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_entregas_cadastrado ON entregas(cadastrado_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_entregas_comb ON entregas(combustivel)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_entregas_posto ON entregas(posto)")

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_cadastrados_nome ON cadastrados(nome)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_cadastrados_cat ON cadastrados(categoria)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_cadastrados_equipe ON cadastrados(equipe)")

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_data ON auditoria(datahora)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_acao ON auditoria(acao)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_entidade ON auditoria(entidade)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_user ON auditoria(username)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_nome_exibicao ON auditoria(nome_exibicao)")

        conn.commit()


def criar_usuarios_padrao():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM usuarios")
        total = cursor.fetchone()[0]

        if total == 0:
            cursor.execute("""
                INSERT INTO usuarios (username, senha, role, nome_exibicao, ativo, criado_em)
                VALUES (?, ?, ?, ?, 1, ?)
            """, ("admin", generate_password_hash("admin123"), "admin", "Administrador", agora_iso()))

            cursor.execute("""
                INSERT INTO usuarios (username, senha, role, nome_exibicao, ativo, criado_em)
                VALUES (?, ?, ?, ?, 1, ?)
            """, ("usuario", generate_password_hash("usuario123"), "user", "Usuário", agora_iso()))

        conn.commit()


def listar_categorias(cursor):
    cursor.execute("SELECT nome FROM categorias ORDER BY nome")
    return [r[0] for r in cursor.fetchall()]


def listar_equipes(cursor):
    cursor.execute("SELECT nome FROM equipes ORDER BY nome")
    return [r[0] for r in cursor.fetchall()]


def get_total_mes(cursor, pessoa_id: int, mes: str) -> float:
    cursor.execute("""
        SELECT IFNULL(SUM(quantidade), 0)
        FROM entregas
        WHERE cadastrado_id = ?
          AND strftime('%Y-%m', data) = ?
    """, (pessoa_id, mes))
    return float(cursor.fetchone()[0] or 0)


def get_total_mes_combustivel(cursor, mes: str, combustivel: str) -> float:
    cursor.execute("""
        SELECT IFNULL(SUM(quantidade), 0)
        FROM entregas
        WHERE strftime('%Y-%m', data) = ?
          AND combustivel = ?
    """, (mes, combustivel))
    return float(cursor.fetchone()[0] or 0)


def get_limite_mensal(cursor, mes: str, combustivel: str) -> float:
    cursor.execute("""
        SELECT IFNULL(limite, 0)
        FROM estoque_mensal
        WHERE mes = ? AND combustivel = ?
    """, (mes, combustivel))
    row = cursor.fetchone()
    return float(row[0] if row else 0)


def upsert_limite_mensal(cursor, mes: str, combustivel: str, limite: float):
    cursor.execute("""
        INSERT INTO estoque_mensal (mes, combustivel, limite)
        VALUES (?, ?, ?)
        ON CONFLICT(mes, combustivel)
        DO UPDATE SET limite = excluded.limite
    """, (mes, combustivel, limite))


def gerar_alerta_estoque(limite: float, usado: float, combustivel: str, mes: str):
    if limite <= 0:
        return None

    pct = (usado / limite) * 100 if limite > 0 else 0
    saldo = limite - usado

    if usado >= limite:
        return {
            "tipo": "vermelho",
            "titulo": f"Limite de {combustivel.upper()} atingido/ultrapassado",
            "msg": f"Mês {mes}: usado {int(usado)}L de {int(limite)}L. Saldo: {int(saldo)}L."
        }

    if pct >= 80:
        return {
            "tipo": "amarelo",
            "titulo": f"{combustivel.upper()} próxima do limite",
            "msg": f"Mês {mes}: usado {int(usado)}L de {int(limite)}L ({pct:.0f}%). Saldo: {int(saldo)}L."
        }

    return None


criar_banco()
criar_usuarios_padrao()


# =========================================================
# AUTH / PERMISSÕES
# =========================================================
def usuario_logado():
    return session.get("user")


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not usuario_logado():
            return redirect("/login?" + urlencode({"next": request.full_path or request.path}))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = usuario_logado()
        if not u:
            return redirect("/login?" + urlencode({"next": request.full_path or request.path}))
        if u.get("role") != "admin":
            abort(403)
        return fn(*args, **kwargs)
    return wrapper


# =========================================================
# LOGIN / LOGOUT / TEMA (ANTI BRUTE-FORCE)
# =========================================================
def _get_ip():
    return (request.headers.get("X-Forwarded-For") or request.remote_addr or "").split(",")[0].strip()


def _login_key(username: str) -> str:
    return f"{_get_ip()}|{(username or '').strip().lower()}"


def _login_fail_register(key: str, max_attempts: int = 7, block_minutes: int = 10):
    store = session.get("login_fail") or {}
    now = datetime.now().timestamp()

    data = store.get(key) or {"count": 0, "blocked_until": 0}
    if data.get("blocked_until", 0) and now < float(data["blocked_until"]):
        store[key] = data
        session["login_fail"] = store
        return True, int(data["blocked_until"] - now)

    data["count"] = int(data.get("count", 0)) + 1

    if data["count"] >= max_attempts:
        data["blocked_until"] = now + (block_minutes * 60)

    store[key] = data
    session["login_fail"] = store

    blocked = now < float(data.get("blocked_until", 0) or 0)
    remaining = int(float(data.get("blocked_until", 0) or 0) - now) if blocked else 0
    return blocked, remaining


def _login_fail_clear(key: str):
    store = session.get("login_fail") or {}
    if key in store:
        store.pop(key, None)
        session["login_fail"] = store


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip().lower()
        senha = request.form.get("senha") or ""
        next_url = (request.form.get("next") or request.args.get("next") or "").strip()

        key = _login_key(username)

        # se estiver bloqueado, avisa
        store = session.get("login_fail") or {}
        data = store.get(key) or {}
        now = datetime.now().timestamp()
        if data.get("blocked_until", 0) and now < float(data["blocked_until"]):
            secs = int(float(data["blocked_until"]) - now)
            mins = max(1, int((secs + 59) / 60))
            flash(f"Muitas tentativas. Aguarde {mins} minuto(s) e tente novamente.", "warning")
            log_acao("LOGIN_BLOQUEADO_TEMP", "usuarios", None, {"username": username, "ip": _get_ip()})
            return render_template("login.html")

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, username, senha, role, nome_exibicao, IFNULL(ativo,1)
            FROM usuarios
            WHERE username = ?
        """, (username,))
        row = cursor.fetchone()
        conn.close()

        if (not row) or (not check_password_hash(row[2], senha)):
            blocked, remaining = _login_fail_register(key)
            if blocked:
                mins = max(1, int((remaining + 59) / 60))
                flash(f"Muitas tentativas. Aguarde {mins} minuto(s) e tente novamente.", "warning")
                log_acao("LOGIN_BLOQUEADO_TEMP", "usuarios", None, {"username": username, "ip": _get_ip()})
            else:
                flash("Usuário ou senha inválidos.", "danger")
                log_acao("LOGIN_FALHA", "usuarios", None, {"username": username, "ip": _get_ip()})
            return render_template("login.html")

        if int(row[5]) == 0:
            flash("Usuário desativado. Fale com o administrador.", "warning")
            log_acao("LOGIN_BLOQUEADO", "usuarios", row[0], {"username": row[1], "motivo": "desativado", "ip": _get_ip()})
            return render_template("login.html")

        _login_fail_clear(key)

        session["user"] = {
            "id": row[0],
            "username": row[1],
            "role": row[3],
            "nome_exibicao": (row[4] or row[1])
        }
        session.setdefault("theme", "dark")

        log_acao("LOGIN_OK", "usuarios", row[0], {"username": row[1], "role": row[3], "ip": _get_ip()})

        # redireciona para next se for relativo (segurança), senão dashboard
        if next_url and next_url.startswith("/"):
            return redirect(next_url)

        return redirect("/dashboard")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    u = usuario_logado() or {}
    log_acao("LOGOUT", "usuarios", u.get("id"), {"username": u.get("username")})

    theme = session.get("theme", "dark")
    session.clear()
    session["theme"] = theme  # mantém tema (opcional)

    return redirect("/login")


# ✅ IMPORTANTE: SEM login_required (pra funcionar na tela de login)
@app.route("/tema/toggle")
def toggle_tema():
    atual = session.get("theme", "dark")
    session["theme"] = "light" if atual == "dark" else "dark"
    return redirect(request.referrer or "/login")


# =========================================================
# CONSULTAS
# =========================================================
def buscar_resumo(mes: str, nome_f: str, categoria_f: str, equipe_f: str, somente_com_movimento: bool = True):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    sql = """
        SELECT
            c.id,
            c.nome,
            c.categoria,
            c.equipe,
            IFNULL(c.cota, 0) as cota,

            IFNULL(SUM(CASE WHEN e.combustivel = 'gasolina' THEN e.quantidade ELSE 0 END), 0) as total_gasolina,
            IFNULL(SUM(CASE WHEN e.combustivel = 'diesel' THEN e.quantidade ELSE 0 END), 0) as total_diesel,

            IFNULL(SUM(e.quantidade), 0) as total_entregue,
            IFNULL(c.cota, 0) - IFNULL(SUM(e.quantidade), 0) as saldo
        FROM cadastrados c
        LEFT JOIN entregas e
            ON c.id = e.cadastrado_id
           AND strftime('%Y-%m', e.data) = ?
        WHERE 1=1
    """
    params = [mes]

    if nome_f:
        sql += " AND c.nome LIKE ?"
        params.append(f"%{nome_f}%")
    if categoria_f:
        sql += " AND c.categoria = ?"
        params.append(categoria_f)
    if equipe_f:
        sql += " AND c.equipe = ?"
        params.append(equipe_f)

    sql += " GROUP BY c.id"

    if somente_com_movimento:
        sql += " HAVING IFNULL(SUM(e.quantidade), 0) > 0"

    sql += " ORDER BY c.nome"

    cursor.execute(sql, tuple(params))
    dados = cursor.fetchall()

    # -------------------------
    # Totais por combustível (como já era)
    # -------------------------
    sql_totais = """
        SELECT e.combustivel, IFNULL(SUM(e.quantidade), 0)
        FROM entregas e
        JOIN cadastrados c ON e.cadastrado_id = c.id
        WHERE strftime('%Y-%m', e.data) = ?
    """
    params_totais = [mes]
    if nome_f:
        sql_totais += " AND c.nome LIKE ?"
        params_totais.append(f"%{nome_f}%")
    if categoria_f:
        sql_totais += " AND c.categoria = ?"
        params_totais.append(categoria_f)
    if equipe_f:
        sql_totais += " AND c.equipe = ?"
        params_totais.append(equipe_f)

    sql_totais += " GROUP BY e.combustivel"
    cursor.execute(sql_totais, tuple(params_totais))
    rows = cursor.fetchall()

    totais_combustivel = {}
    for r in rows:
        # caso normal: (combustivel, total)
        if len(r) == 2:
            comb, total = r
            totais_combustivel[(comb or "").lower()] = float(total or 0)

        # caso você tenha posto junto: (posto, combustivel, total)
        elif len(r) >= 3:
            posto, comb, total = r[0], r[1], r[2]
            # aqui você decide: somar tudo geral por combustível
            key = (comb or "").lower()
            totais_combustivel[key] = totais_combustivel.get(key, 0) + float(total or 0)

        # garante chaves
        totais_combustivel.setdefault("gasolina", 0.0)
        totais_combustivel.setdefault("diesel", 0.0)

    # -------------------------
    # NOVO: Totais por posto (PONTAO / ESTRADA / CIDADE + TOTAL)
    # -------------------------
    sql_postos = """
        SELECT IFNULL(e.posto,'CIDADE') as posto, IFNULL(SUM(e.quantidade), 0)
        FROM entregas e
        JOIN cadastrados c ON e.cadastrado_id = c.id
        WHERE strftime('%Y-%m', e.data) = ?
    """
    params_postos = [mes]
    if nome_f:
        sql_postos += " AND c.nome LIKE ?"
        params_postos.append(f"%{nome_f}%")
    if categoria_f:
        sql_postos += " AND c.categoria = ?"
        params_postos.append(categoria_f)
    if equipe_f:
        sql_postos += " AND c.equipe = ?"
        params_postos.append(equipe_f)

    sql_postos += " GROUP BY IFNULL(e.posto,'CIDADE')"
    cursor.execute(sql_postos, tuple(params_postos))
    rows_postos = cursor.fetchall()

    totais_posto = { (k or "CIDADE").upper(): float(v or 0) for k, v in rows_postos }
    # normaliza chaves esperadas
    if "PONTÃO" in totais_posto and "PONTAO" not in totais_posto:
        totais_posto["PONTAO"] = totais_posto.pop("PONTÃO")
    totais_posto.setdefault("PONTAO", 0.0)
    totais_posto.setdefault("ESTRADA", 0.0)
    totais_posto.setdefault("CIDADE", 0.0)
    totais_posto["TOTAL"] = float(totais_posto.get("PONTAO", 0.0) + totais_posto.get("ESTRADA", 0.0) + totais_posto.get("CIDADE", 0.0))

    conn.close()
    return dados, totais_combustivel, totais_posto


# =========================================================
# CONSULTAS
# =========================================================

POSTOS_VALIDOS = {"PONTAO", "ESTRADA", "CIDADE"}

def _normalizar_posto_db(posto: str) -> str:
    p = (posto or "CIDADE").strip().upper()
    if p == "PONTÃO":
        p = "PONTAO"
    if p not in POSTOS_VALIDOS:
        p = "CIDADE"
    return p


def buscar_entregas(mes: str, nome_f: str, categoria_f: str, equipe_f: str):
    mes = (mes or "").strip()
    nome_f = (nome_f or "").strip()
    categoria_f = (categoria_f or "").strip()
    equipe_f = (equipe_f or "").strip()

    where = []
    params = []

    if mes:
        where.append("strftime('%Y-%m', e.data) = ?")
        params.append(mes)

    if nome_f:
        where.append("UPPER(c.nome) LIKE UPPER(?)")
        params.append(f"%{nome_f}%")

    if categoria_f:
        where.append("c.categoria = ?")
        params.append(categoria_f)

    if equipe_f:
        where.append("c.equipe = ?")
        params.append(equipe_f)

    where_sql = (" AND " + " AND ".join(where)) if where else ""

    sql_list = f"""
        SELECT
            e.id, e.cadastrado_id, e.quantidade, e.data, e.observacao, e.combustivel,
            c.nome, c.categoria, c.equipe,
            IFNULL(e.posto,'CIDADE') as posto
        FROM entregas e
        JOIN cadastrados c ON e.cadastrado_id = c.id
        WHERE 1=1
        {where_sql}
        ORDER BY e.data DESC, e.id DESC
    """

    sql_tot = f"""
        SELECT e.combustivel, IFNULL(SUM(e.quantidade),0)
        FROM entregas e
        JOIN cadastrados c ON e.cadastrado_id = c.id
        WHERE 1=1
        {where_sql}
        GROUP BY e.combustivel
    """

    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute(sql_list, tuple(params))
        entregas_raw = cursor.fetchall()

        cursor.execute(sql_tot, tuple(params))
        rows = cursor.fetchall()

    # garante posto sempre OK (PONTAO/ESTRADA/CIDADE)
    entregas = []
    for e in entregas_raw:
        e = list(e)
        e[9] = _normalizar_posto_db(e[9])
        entregas.append(tuple(e))

    totais = {k: float(v or 0) for k, v in rows}
    totais.setdefault("gasolina", 0.0)
    totais.setdefault("diesel", 0.0)

    return entregas, totais


def buscar_entregas_paginado(mes, nome_f, categoria_f, equipe_f, page=1, per_page=25):
    try:
        page = int(page or 1)
    except:
        page = 1
    if page < 1:
        page = 1

    try:
        per_page = int(per_page or 25)
    except:
        per_page = 25
    if per_page < 1:
        per_page = 25

    offset = (page - 1) * per_page

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    where = []
    params = []

    if mes:
        where.append("substr(e.data, 1, 7) = ?")
        params.append(mes)

    if nome_f:
        # deixa igual ao buscar_entregas (case-insensitive)
        where.append("UPPER(c.nome) LIKE UPPER(?)")
        params.append(f"%{nome_f}%")

    if categoria_f:
        where.append("IFNULL(c.categoria,'') = ?")
        params.append(categoria_f)

    if equipe_f:
        where.append("IFNULL(c.equipe,'') = ?")
        params.append(equipe_f)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    try:
        cursor.execute(f"""
            SELECT COUNT(*)
            FROM entregas e
            JOIN cadastrados c ON c.id = e.cadastrado_id
            {where_sql}
        """, params)
        total_registros = int((cursor.fetchone() or [0])[0] or 0)

        cursor.execute(f"""
            SELECT
                e.id,
                e.cadastrado_id,
                e.quantidade,
                e.data,
                e.observacao,
                e.combustivel,
                c.nome,
                IFNULL(c.categoria,''),
                IFNULL(c.equipe,''),
                IFNULL(e.posto,'CIDADE') as posto
            FROM entregas e
            JOIN cadastrados c ON c.id = e.cadastrado_id
            {where_sql}
            ORDER BY e.data DESC, e.id DESC
            LIMIT ? OFFSET ?
        """, params + [per_page, offset])
        entregas_raw = cursor.fetchall()

        # normaliza posto sempre
        entregas = []
        for e in entregas_raw:
            e = list(e)
            e[9] = _normalizar_posto_db(e[9])
            entregas.append(tuple(e))

        cursor.execute(f"""
            SELECT e.combustivel, COALESCE(SUM(e.quantidade), 0)
            FROM entregas e
            JOIN cadastrados c ON c.id = e.cadastrado_id
            {where_sql}
            GROUP BY e.combustivel
        """, params)
        rows = cursor.fetchall()

        totais_filtrados = {(r[0] or "").lower(): float(r[1] or 0) for r in rows}
        totais_filtrados.setdefault("gasolina", 0.0)
        totais_filtrados.setdefault("diesel", 0.0)

        return entregas, totais_filtrados, total_registros

    finally:
        conn.close()

# =========================================================
# ROTAS BÁSICAS / DASHBOARD
# =========================================================
@app.route("/")
def index():
    u = usuario_logado()
    if not u:
        return redirect("/login")
    return redirect("/dashboard")


@app.route("/dashboard")
@login_required
def dashboard():
    mes = request.args.get("mes") or datetime.now().strftime("%Y-%m")

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    limite_gas = get_limite_mensal(cursor, mes, "gasolina")
    limite_diesel = get_limite_mensal(cursor, mes, "diesel")

    usado_gas = get_total_mes_combustivel(cursor, mes, "gasolina")
    usado_diesel = get_total_mes_combustivel(cursor, mes, "diesel")

    saldo_gas = limite_gas - usado_gas
    saldo_diesel = limite_diesel - usado_diesel

    alerta_gas = gerar_alerta_estoque(limite_gas, usado_gas, "gasolina", mes)
    alerta_diesel = gerar_alerta_estoque(limite_diesel, usado_diesel, "diesel", mes)

    cursor.execute("""
        SELECT c.id, c.nome, c.categoria, c.equipe, IFNULL(SUM(e.quantidade),0) as total
        FROM cadastrados c
        LEFT JOIN entregas e
               ON e.cadastrado_id = c.id
              AND strftime('%Y-%m', e.data) = ?
        GROUP BY c.id
        ORDER BY total DESC, c.nome ASC
        LIMIT 5
    """, (mes,))
    top5 = cursor.fetchall()

    cursor.execute("""
        SELECT IFNULL(SUM(quantidade),0)
        FROM entregas
        WHERE strftime('%Y-%m', data) = ?
    """, (mes,))
    total_geral = float(cursor.fetchone()[0] or 0)

    cursor.execute("""
        SELECT strftime('%d', data) as dia, IFNULL(SUM(quantidade),0) as total
        FROM entregas
        WHERE strftime('%Y-%m', data) = ?
        GROUP BY strftime('%Y-%m-%d', data)
        ORDER BY strftime('%Y-%m-%d', data) ASC
    """, (mes,))
    serie = cursor.fetchall()

    serie_dias = [row[0] for row in serie]
    serie_vals = [float(row[1] or 0) for row in serie]

    conn.close()

    def pct(usado, limite):
        if limite and limite > 0:
            return round((usado / limite) * 100, 0)
        return None

    return render_template(
        "dashboard.html",
        mes=mes,
        total_geral=int(total_geral),
        usado_gas=int(usado_gas),
        usado_diesel=int(usado_diesel),
        limite_gas=int(limite_gas),
        limite_diesel=int(limite_diesel),
        saldo_gas=int(saldo_gas),
        saldo_diesel=int(saldo_diesel),
        pct_gas=pct(usado_gas, limite_gas),
        pct_diesel=pct(usado_diesel, limite_diesel),
        alerta_gas=alerta_gas,
        alerta_diesel=alerta_diesel,
        top5=top5,
        serie_dias=serie_dias,
        serie_vals=serie_vals,
        u=usuario_logado()
    )


# =========================================================
# AUDITORIA (ADMIN)
# =========================================================
@app.route("/auditoria")
@admin_required
def auditoria():
    q = (request.args.get("q") or "").strip()
    acao = (request.args.get("acao") or "").strip()
    entidade = (request.args.get("entidade") or "").strip()
    username = (request.args.get("username") or "").strip()
    data_ini = (request.args.get("ini") or "").strip()
    data_fim = (request.args.get("fim") or "").strip()

    sql = """
        SELECT id, datahora, user_id, username, nome_exibicao, role,
               acao, entidade, entidade_id, detalhes, ip
        FROM auditoria
        WHERE 1=1
    """
    params = []

    if q:
        like = f"%{q}%"
        sql += " AND (acao LIKE ? OR entidade LIKE ? OR detalhes LIKE ? OR username LIKE ? OR nome_exibicao LIKE ?)"
        params += [like, like, like, like, like]

    if acao:
        sql += " AND acao = ?"
        params.append(acao)

    if entidade:
        sql += " AND entidade = ?"
        params.append(entidade)

    if username:
        likeu = f"%{username}%"
        sql += " AND (username LIKE ? OR nome_exibicao LIKE ?)"
        params += [likeu, likeu]

    if data_ini:
        sql += " AND substr(datahora,1,10) >= ?"
        params.append(data_ini)

    if data_fim:
        sql += " AND substr(datahora,1,10) <= ?"
        params.append(data_fim)

    sql += " ORDER BY id DESC LIMIT 300"

    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()

        cursor.execute("SELECT DISTINCT acao FROM auditoria ORDER BY acao")
        acoes = [r[0] for r in cursor.fetchall()]

        cursor.execute("""
            SELECT DISTINCT entidade
            FROM auditoria
            WHERE entidade IS NOT NULL AND entidade <> ''
            ORDER BY entidade
        """)
        entidades = [r[0] for r in cursor.fetchall()]

    return render_template(
        "auditoria.html",
        rows=rows,
        acoes=acoes,
        entidades=entidades,
        q=q,
        acao_sel=acao,
        entidade_sel=entidade,
        username_sel=username,
        ini=data_ini,
        fim=data_fim,
        u=session.get("user")
    )


@app.route("/auditoria/apagar-tudo", methods=["POST"])
@admin_required
def auditoria_apagar_tudo():
    u = usuario_atual() or {}
    ip = request.headers.get("X-Forwarded-For", request.remote_addr)

    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(*) FROM auditoria")
            total = int(cursor.fetchone()[0] or 0)

            log_acao("LIMPAR_AUDITORIA", "auditoria", None, {"apagados": total})

            detalhes = json.dumps({"apagados": total}, ensure_ascii=False)
            cursor.execute("""
                INSERT INTO auditoria_admin (datahora, user_id, username, nome_exibicao, role, acao, detalhes, ip)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                agora_iso(),
                u.get("id"),
                u.get("username"),
                u.get("nome_exibicao"),
                u.get("role"),
                "LIMPAR_AUDITORIA",
                detalhes,
                ip
            ))

            cursor.execute("DELETE FROM auditoria")

        flash(f"Auditoria limpa com sucesso. ({total} registros apagados)", "success")
        return redirect("/auditoria")

    except Exception as e:
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO auditoria_admin (datahora, user_id, username, nome_exibicao, role, acao, detalhes, ip)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    agora_iso(),
                    u.get("id"),
                    u.get("username"),
                    u.get("nome_exibicao"),
                    u.get("role"),
                    "LIMPAR_AUDITORIA_FALHA",
                    json.dumps({"erro": str(e)}, ensure_ascii=False),
                    ip
                ))
        except:
            pass

        flash("Não foi possível limpar a auditoria.", "danger")
        return redirect("/auditoria")


# =========================================================
# CADASTRO (ADMIN)
# =========================================================
@app.route("/cadastro", methods=["GET", "POST"])
@login_required
def cadastro():
    u = usuario_logado()
    if u.get("role") != "admin":
        return redirect("/entrega")

    nome_f = (request.args.get("nome") or "").strip()
    categoria_f = normalizar(request.args.get("categoria") or "")
    equipe_f = normalizar(request.args.get("equipe") or "")

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        categorias_lista = listar_categorias(cursor)
        equipes_lista = listar_equipes(cursor)

        if request.method == "POST":
            nome = (request.form.get("nome") or "").strip()
            categoria = normalizar(request.form.get("categoria") or "")
            equipe = normalizar(request.form.get("equipe") or "")

            cota_raw = (request.form.get("cota") or "").strip().replace(",", ".")
            try:
                cota = float(cota_raw) if cota_raw else 0.0
            except ValueError:
                flash("Cota inválida. Use apenas números (ex: 50 ou 50,5).", "warning")
                return redirect("/cadastro")

            if not nome:
                flash("Informe o nome.", "warning")
                return redirect("/cadastro")

            cursor.execute(
                "INSERT INTO cadastrados (nome, categoria, equipe, cota) VALUES (?, ?, ?, ?)",
                (nome, categoria, equipe, cota)
            )
            pessoa_id = cursor.lastrowid
            conn.commit()

            log_acao("CRIAR_CADASTRO", "cadastrados", pessoa_id, {
                "nome": nome, "categoria": categoria, "equipe": equipe, "cota": cota
            })

            flash("Cadastro criado com sucesso.", "success")
            return redirect("/cadastro")

        sql = "SELECT id, nome, categoria, equipe, cota FROM cadastrados WHERE 1=1"
        params = []

        if nome_f:
            sql += " AND nome LIKE ?"
            params.append(f"%{nome_f}%")
        if categoria_f:
            sql += " AND categoria = ?"
            params.append(categoria_f)
        if equipe_f:
            sql += " AND equipe = ?"
            params.append(equipe_f)

        sql += " ORDER BY nome"
        cursor.execute(sql, tuple(params))
        pessoas = cursor.fetchall()

        return render_template(
            "cadastro.html",
            pessoas=pessoas,
            filtro_nome=nome_f,
            filtro_categoria=categoria_f,
            filtro_equipe=equipe_f,
            categorias_lista=categorias_lista,
            equipes_lista=equipes_lista
        )

    finally:
        conn.close()


@app.route("/cadastro/apagar/<int:pessoa_id>", methods=["POST"])
@admin_required
def apagar_cadastro(pessoa_id):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT nome, categoria, equipe, IFNULL(cota,0) FROM cadastrados WHERE id=?", (pessoa_id,))
        row = cursor.fetchone()
        if not row:
            flash("Cadastro não encontrado.", "warning")
            return redirect("/cadastro")

        nome, categoria, equipe, cota = row[0], row[1], row[2], float(row[3] or 0)

        cursor.execute("DELETE FROM entregas WHERE cadastrado_id = ?", (pessoa_id,))
        cursor.execute("DELETE FROM cadastrados WHERE id = ?", (pessoa_id,))
        conn.commit()

        log_acao("APAGAR_CADASTRO", "cadastrados", pessoa_id, {
            "nome": nome,
            "categoria": categoria,
            "equipe": equipe,
            "cota": cota
        })

        flash("Cadastro apagado com sucesso.", "success")
        return redirect("/cadastro")

    finally:
        conn.close()


@app.route("/cadastro/editar/<int:pessoa_id>", methods=["GET", "POST"])
@admin_required
def editar_cadastro(pessoa_id):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        categorias_lista = listar_categorias(cursor)
        equipes_lista = listar_equipes(cursor)

        if request.method == "POST":
            nome = (request.form.get("nome") or "").strip()
            categoria = normalizar(request.form.get("categoria") or "")
            equipe = normalizar(request.form.get("equipe") or "")

            cota_raw = (request.form.get("cota") or "").strip().replace(",", ".")
            try:
                cota = float(cota_raw) if cota_raw else 0.0
            except ValueError:
                flash("Cota inválida. Use apenas números (ex: 50 ou 50,5).", "warning")
                return redirect(f"/cadastro/editar/{pessoa_id}")

            if not nome:
                flash("Informe o nome.", "warning")
                return redirect(f"/cadastro/editar/{pessoa_id}")

            cursor.execute("""
                UPDATE cadastrados
                   SET nome = ?, categoria = ?, equipe = ?, cota = ?
                 WHERE id = ?
            """, (nome, categoria, equipe, cota, pessoa_id))
            conn.commit()

            log_acao("EDITAR_CADASTRO", "cadastrados", pessoa_id, {
                "nome": nome, "categoria": categoria, "equipe": equipe, "cota": cota
            })

            flash("Cadastro atualizado com sucesso.", "success")
            return redirect("/cadastro")

        cursor.execute("SELECT id, nome, categoria, equipe, IFNULL(cota,0) FROM cadastrados WHERE id = ?", (pessoa_id,))
        pessoa = cursor.fetchone()
        if not pessoa:
            flash("Cadastro não encontrado.", "warning")
            return redirect("/cadastro")

        return render_template(
            "editar_cadastro.html",
            p=pessoa,
            categorias_lista=categorias_lista,
            equipes_lista=equipes_lista
        )

    finally:
        conn.close()


# =========================================================
# ENTREGA (ADMIN E USER)
# =========================================================

POSTOS_VALIDOS = {"PONTAO", "ESTRADA", "CIDADE"}

@app.route("/entrega", methods=["GET", "POST"])
@login_required
def entrega():
    mes = normalizar_mes_param(request.args.get("mes")) or ""
    nome_f = (request.args.get("nome") or "").strip()
    categoria_f = normalizar(request.args.get("categoria") or "")
    equipe_f = normalizar(request.args.get("equipe") or "")

    # PAGINAÇÃO
    try:
        page = int(request.args.get("page") or 1)
    except:
        page = 1
    if page < 1:
        page = 1

    try:
        per_page = int(request.args.get("per_page") or 25)
    except:
        per_page = 25
    if per_page not in (10, 25, 50, 100):
        per_page = 25

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        categorias_lista = listar_categorias(cursor)
        equipes_lista = listar_equipes(cursor)

        cursor.execute("SELECT id, nome FROM cadastrados ORDER BY nome")
        pessoas = cursor.fetchall()

        if request.method == "POST":
            # pega filtros do form (pra manter estado mesmo após salvar)
            mes_form = normalizar_mes_param(request.form.get("mes_filtro")) or mes
            nome_form = (request.form.get("nome_filtro") or nome_f).strip()
            categoria_form = normalizar(request.form.get("categoria_filtro") or categoria_f)
            equipe_form = normalizar(request.form.get("equipe_filtro") or equipe_f)

            try:
                pessoa_id = int(request.form.get("pessoa_id") or 0)
            except:
                pessoa_id = 0

            data = (request.form.get("data") or "").strip()
            observacao = (request.form.get("observacao") or "").strip()
            combustivel = (request.form.get("combustivel") or "").strip().lower()

            # ======================
            # POSTO
            # ======================
            posto = (request.form.get("posto") or "CIDADE").strip().upper()
            if posto == "PONTÃO":
                posto = "PONTAO"
            if posto not in POSTOS_VALIDOS:
                flash("Posto inválido. Escolha Pontão, Estrada ou Cidade.", "warning")
                # volta mantendo filtros
                params = {"mes": mes_form or ""}
                if nome_form: params["nome"] = nome_form
                if categoria_form: params["categoria"] = categoria_form
                if equipe_form: params["equipe"] = equipe_form
                params["per_page"] = per_page
                params["page"] = page
                return redirect("/entrega?" + urlencode(params))

            try:
                quantidade = float(request.form.get("quantidade") or 0)
            except:
                quantidade = 0

            if pessoa_id <= 0:
                flash("Selecione a pessoa/setor.", "warning")
                params = {"mes": mes_form or ""}
                if nome_form: params["nome"] = nome_form
                if categoria_form: params["categoria"] = categoria_form
                if equipe_form: params["equipe"] = equipe_form
                params["per_page"] = per_page
                params["page"] = page
                return redirect("/entrega?" + urlencode(params))

            cursor.execute("SELECT id, nome, IFNULL(cota,0) FROM cadastrados WHERE id=?", (pessoa_id,))
            pessoa = cursor.fetchone()
            if not pessoa:
                flash("Cadastro não encontrado.", "danger")
                return redirect("/entrega")

            if not re.match(r"^\d{4}-\d{2}-\d{2}$", data):
                flash("Data inválida.", "warning")
                params = {"mes": mes_form or ""}
                if nome_form: params["nome"] = nome_form
                if categoria_form: params["categoria"] = categoria_form
                if equipe_form: params["equipe"] = equipe_form
                params["per_page"] = per_page
                params["page"] = page
                return redirect("/entrega?" + urlencode(params))

            if quantidade <= 0:
                flash("Quantidade deve ser maior que zero.", "warning")
                params = {"mes": data[:7]}
                if nome_form: params["nome"] = nome_form
                if categoria_form: params["categoria"] = categoria_form
                if equipe_form: params["equipe"] = equipe_form
                params["per_page"] = per_page
                params["page"] = page
                return redirect("/entrega?" + urlencode(params))

            if combustivel not in ("gasolina", "diesel"):
                flash("Combustível inválido.", "danger")
                params = {"mes": data[:7]}
                if nome_form: params["nome"] = nome_form
                if categoria_form: params["categoria"] = categoria_form
                if equipe_form: params["equipe"] = equipe_form
                params["per_page"] = per_page
                params["page"] = page
                return redirect("/entrega?" + urlencode(params))

            cursor.execute("""
                INSERT INTO entregas
                (cadastrado_id, quantidade, data, observacao, combustivel, posto)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (pessoa_id, quantidade, data, observacao, combustivel, posto))

            entrega_id = cursor.lastrowid
            conn.commit()

            log_acao("CRIAR_ENTREGA", "entregas", entrega_id, {
                "pessoa_id": pessoa_id,
                "data": data,
                "quantidade": quantidade,
                "combustivel": combustivel,
                "posto": posto,
                "observacao": observacao
            })

            flash("Entrega registrada com sucesso.", "success")

            # volta no mês da data registrada, mantendo filtros do form
            params = {"mes": data[:7]}
            if nome_form:
                params["nome"] = nome_form
            if categoria_form:
                params["categoria"] = categoria_form
            if equipe_form:
                params["equipe"] = equipe_form
            params["per_page"] = per_page
            params["page"] = page

            return redirect("/entrega?" + urlencode(params))

        entregas, totais_filtrados, total_registros = buscar_entregas_paginado(
            mes, nome_f, categoria_f, equipe_f, page=page, per_page=per_page
        )

        total_pages = max(1, (total_registros + per_page - 1) // per_page)
        if page > total_pages:
            page = total_pages

        return render_template(
            "entrega.html",
            pessoas=pessoas,
            entregas=entregas,
            totais_filtrados=totais_filtrados,
            mes_selecionado=mes,
            filtro_nome=nome_f,
            filtro_categoria=categoria_f,
            filtro_equipe=equipe_f,
            categorias_lista=categorias_lista,
            equipes_lista=equipes_lista,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            total_registros=total_registros
        )

    finally:
        conn.close()

@app.route("/entrega/apagar/<int:entrega_id>", methods=["POST"])
@admin_required
def apagar_entrega(entrega_id):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT cadastrado_id, data, quantidade, combustivel, observacao, IFNULL(posto,'CIDADE')
            FROM entregas
            WHERE id=?
        """, (entrega_id,))
        row = cursor.fetchone()

        detalhes = {}
        if row:
            detalhes = {
                "pessoa_id": row[0],
                "data": row[1],
                "quantidade": row[2],
                "combustivel": row[3],
                "observacao": row[4],
                "posto": row[5]
            }

        cursor.execute("DELETE FROM entregas WHERE id=?", (entrega_id,))
        conn.commit()

        log_acao("APAGAR_ENTREGA", "entregas", entrega_id, detalhes)
        flash("Entrega apagada.", "success")

    finally:
        conn.close()

    next_url = request.form.get("next") or request.args.get("next") or "/entrega"
    return redirect(next_url)


@app.route("/entrega/editar/<int:entrega_id>", methods=["GET", "POST"])
@admin_required
def editar_entrega(entrega_id):
    POSTOS_VALIDOS = {"CIDADE", "ESTRADA", "PONTAO"}

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        if request.method == "POST":
            try:
                pessoa_id = int(request.form.get("pessoa_id") or 0)
            except:
                pessoa_id = 0

            data = (request.form.get("data") or "").strip()

            try:
                quantidade = float(request.form.get("quantidade") or 0)
            except:
                quantidade = 0

            observacao = (request.form.get("observacao") or "").strip()
            combustivel = (request.form.get("combustivel") or "").strip().lower()

            # NOVO: POSTO
            posto = (request.form.get("posto") or "CIDADE").strip().upper()
            if posto == "PONTÃO":
                posto = "PONTAO"
            if posto not in POSTOS_VALIDOS:
                flash("Posto inválido.", "warning")
                return redirect(f"/entrega/editar/{entrega_id}")

            if pessoa_id <= 0:
                flash("Selecione a pessoa/setor.", "warning")
                return redirect(f"/entrega/editar/{entrega_id}")

            if not re.match(r"^\d{4}-\d{2}-\d{2}$", data):
                flash("Data inválida.", "warning")
                return redirect(f"/entrega/editar/{entrega_id}")

            if quantidade <= 0:
                flash("Quantidade deve ser maior que zero.", "warning")
                return redirect(f"/entrega/editar/{entrega_id}")

            if combustivel not in ("gasolina", "diesel"):
                flash("Combustível inválido.", "danger")
                return redirect(f"/entrega/editar/{entrega_id}")

            cursor.execute("""
                UPDATE entregas
                SET cadastrado_id=?, quantidade=?, data=?, observacao=?, combustivel=?, posto=?
                WHERE id=?
            """, (pessoa_id, quantidade, data, observacao, combustivel, posto, entrega_id))

            conn.commit()

            log_acao("EDITAR_ENTREGA", "entregas", entrega_id, {
                "pessoa_id": pessoa_id,
                "data": data,
                "quantidade": quantidade,
                "combustivel": combustivel,
                "posto": posto,
                "observacao": observacao
            })

            flash("Entrega atualizada.", "success")
            return redirect("/entrega")

        cursor.execute("SELECT * FROM entregas WHERE id=?", (entrega_id,))
        entrega = cursor.fetchone()
        if not entrega:
            flash("Entrega não encontrada.", "warning")
            return redirect("/entrega")

        cursor.execute("SELECT id, nome FROM cadastrados ORDER BY nome")
        pessoas = cursor.fetchall()

        return render_template("editar_entrega.html", entrega=entrega, pessoas=pessoas)

    finally:
        conn.close()

# =========================================================
# CONSULTAS
# =========================================================
def buscar_resumo(mes: str, nome_f: str, categoria_f: str, equipe_f: str, somente_com_movimento: bool = True):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    sql = """
        SELECT
            c.id,
            c.nome,
            c.categoria,
            c.equipe,
            IFNULL(c.cota, 0) as cota,

            IFNULL(SUM(CASE WHEN e.combustivel = 'gasolina' THEN e.quantidade ELSE 0 END), 0) as total_gasolina,
            IFNULL(SUM(CASE WHEN e.combustivel = 'diesel' THEN e.quantidade ELSE 0 END), 0) as total_diesel,

            IFNULL(SUM(e.quantidade), 0) as total_entregue,
            IFNULL(c.cota, 0) - IFNULL(SUM(e.quantidade), 0) as saldo
        FROM cadastrados c
        LEFT JOIN entregas e
            ON c.id = e.cadastrado_id
           AND strftime('%Y-%m', e.data) = ?
        WHERE 1=1
    """
    params = [mes]

    if nome_f:
        sql += " AND c.nome LIKE ?"
        params.append(f"%{nome_f}%")
    if categoria_f:
        sql += " AND c.categoria = ?"
        params.append(categoria_f)
    if equipe_f:
        sql += " AND c.equipe = ?"
        params.append(equipe_f)

    sql += " GROUP BY c.id"

    if somente_com_movimento:
        sql += " HAVING IFNULL(SUM(e.quantidade), 0) > 0"

    sql += " ORDER BY c.nome"

    cursor.execute(sql, tuple(params))
    dados = cursor.fetchall()

    # -------------------------
    # Totais por combustível (2 colunas)
    # -------------------------
    sql_totais = """
        SELECT e.combustivel, IFNULL(SUM(e.quantidade), 0)
        FROM entregas e
        JOIN cadastrados c ON e.cadastrado_id = c.id
        WHERE strftime('%Y-%m', e.data) = ?
    """
    params_totais = [mes]
    if nome_f:
        sql_totais += " AND c.nome LIKE ?"
        params_totais.append(f"%{nome_f}%")
    if categoria_f:
        sql_totais += " AND c.categoria = ?"
        params_totais.append(categoria_f)
    if equipe_f:
        sql_totais += " AND c.equipe = ?"
        params_totais.append(equipe_f)

    sql_totais += " GROUP BY e.combustivel"
    cursor.execute(sql_totais, tuple(params_totais))
    rows = cursor.fetchall()

    totais_combustivel = {}
    for comb, total in rows:  # ✅ sempre 2 valores
        totais_combustivel[(comb or "").lower()] = float(total or 0)

    totais_combustivel.setdefault("gasolina", 0.0)
    totais_combustivel.setdefault("diesel", 0.0)

    # -------------------------
    # NOVO: Totais por posto (PONTAO / ESTRADA / CIDADE + TOTAL)
    # -------------------------
    sql_postos = """
        SELECT IFNULL(e.posto,'CIDADE') as posto, IFNULL(SUM(e.quantidade), 0)
        FROM entregas e
        JOIN cadastrados c ON e.cadastrado_id = c.id
        WHERE strftime('%Y-%m', e.data) = ?
    """
    params_postos = [mes]
    if nome_f:
        sql_postos += " AND c.nome LIKE ?"
        params_postos.append(f"%{nome_f}%")
    if categoria_f:
        sql_postos += " AND c.categoria = ?"
        params_postos.append(categoria_f)
    if equipe_f:
        sql_postos += " AND c.equipe = ?"
        params_postos.append(equipe_f)

    sql_postos += " GROUP BY IFNULL(e.posto,'CIDADE')"
    cursor.execute(sql_postos, tuple(params_postos))
    rows_postos = cursor.fetchall()

    totais_posto = {(k or "CIDADE").strip().upper(): float(v or 0) for k, v in rows_postos}
    # normaliza "PONTÃO" -> "PONTAO"
    if "PONTÃO" in totais_posto and "PONTAO" not in totais_posto:
        totais_posto["PONTAO"] = totais_posto.pop("PONTÃO")

    totais_posto.setdefault("PONTAO", 0.0)
    totais_posto.setdefault("ESTRADA", 0.0)
    totais_posto.setdefault("CIDADE", 0.0)
    totais_posto["TOTAL"] = float(
        totais_posto.get("PONTAO", 0.0) +
        totais_posto.get("ESTRADA", 0.0) +
        totais_posto.get("CIDADE", 0.0)
    )

    conn.close()
    return dados, totais_combustivel, totais_posto


# =========================================================
# RESUMO / DETALHE
# =========================================================

@app.route("/resumo")
@login_required
def resumo():
    mes = normalizar_mes_param(request.args.get("mes")) or datetime.now().strftime("%Y-%m")
    nome_f = (request.args.get("nome") or "").strip()
    categoria_f = normalizar(request.args.get("categoria") or "")
    equipe_f = normalizar(request.args.get("equipe") or "")

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    categorias_lista = listar_categorias(cursor)
    equipes_lista = listar_equipes(cursor)
    conn.close()

    # ✅ agora recebe 3 valores
    dados, totais_combustivel, totais_posto = buscar_resumo(
        mes, nome_f, categoria_f, equipe_f, somente_com_movimento=True
    )

    return render_template(
        "resumo.html",
        dados=dados,
        totais_combustivel=totais_combustivel,
        totais_posto=totais_posto,
        mes=mes,
        filtro_nome=nome_f,
        filtro_categoria=categoria_f,
        filtro_equipe=equipe_f,
        categorias_lista=categorias_lista,
        equipes_lista=equipes_lista
    )


@app.route("/detalhe")
@login_required
def detalhe():
    mes = request.args.get("mes") or datetime.now().strftime("%Y-%m")
    pessoa_id = request.args.get("pessoa_id")

    if not pessoa_id:
        return redirect(f"/resumo?mes={mes}")

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute(
        "SELECT id, nome, categoria, equipe, IFNULL(cota,0) FROM cadastrados WHERE id=?",
        (pessoa_id,)
    )
    pessoa = cursor.fetchone()
    if not pessoa:
        conn.close()
        return redirect(f"/resumo?mes={mes}")

    cursor.execute("""
        SELECT id, data, quantidade, combustivel, observacao, IFNULL(posto,'CIDADE')
        FROM entregas
        WHERE cadastrado_id=?
          AND strftime('%Y-%m', data)=?
        ORDER BY data
    """, (pessoa_id, mes))
    entregas = cursor.fetchall()

    cursor.execute("""
        SELECT combustivel, IFNULL(SUM(quantidade),0)
        FROM entregas
        WHERE cadastrado_id=?
          AND strftime('%Y-%m', data)=?
        GROUP BY combustivel
    """, (pessoa_id, mes))
    totais = dict(cursor.fetchall())

    conn.close()

    return render_template(
        "detalhe.html",
        mes=mes,
        pessoa=pessoa,
        entregas=entregas,
        totais=totais
    )


# =========================================================
# PDF — ASSINATURA
# =========================================================
def desenhar_assinatura_pdf(canvas, doc_obj):
    w, h = doc_obj.pagesize
    y = 35
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(1)
    canvas.line(doc_obj.leftMargin, y + 18, w - doc_obj.rightMargin, y + 18)
    canvas.setFont("Helvetica", 9)
    canvas.drawString(
        doc_obj.leftMargin, y + 5,
        f"Assinatura: {RESPONSAVEL_NOME} — {RESPONSAVEL_CARGO} — {RESPONSAVEL_SETOR}"
    )
    canvas.drawRightString(w - doc_obj.rightMargin, y + 5, f"Emissão: {agora_emissao_str()}")


# =========================================================
# EXPORTAÇÕES (RESUMO)
# =========================================================

@app.route("/resumo/excel")
@login_required
def resumo_excel():
    mes = request.args.get("mes") or datetime.now().strftime("%Y-%m")
    nome_f = (request.args.get("nome") or "").strip()
    categoria_f = normalizar(request.args.get("categoria") or "")
    equipe_f = normalizar(request.args.get("equipe") or "")

    # ✅ agora retorna 3
    dados, totais_combustivel, totais_posto = buscar_resumo(
        mes, nome_f, categoria_f, equipe_f, somente_com_movimento=False
    )
    emissao = agora_emissao_str()

    log_acao("EXPORTAR_RESUMO_EXCEL", "resumo", None, {
        "mes": mes, "nome": nome_f, "categoria": categoria_f, "equipe": equipe_f
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    ws["A1"] = f"RESUMO TOTAL MÊS - {mes}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Emissão: {emissao}"
    ws["A3"] = f"Responsável: {RESPONSAVEL_NOME} - {RESPONSAVEL_CARGO} - {RESPONSAVEL_SETOR}"
    ws["A4"] = f"Filtros: nome='{nome_f or 'TODOS'}', categoria='{categoria_f or 'TODOS'}', equipe='{equipe_f or 'TODOS'}'"
    ws["A4"].font = Font(italic=True)

    gasolina_total = int(totais_combustivel.get("gasolina", 0) or 0)
    diesel_total = int(totais_combustivel.get("diesel", 0) or 0)
    total_geral = int(totais_posto.get("TOTAL", gasolina_total + diesel_total) or 0)

    ws["A6"] = f"Gasolina: {gasolina_total}L"
    ws["B6"] = f"Diesel: {diesel_total}L"
    ws["C6"] = f"Total geral: {total_geral}L"
    ws["A6"].font = Font(bold=True)
    ws["B6"].font = Font(bold=True)
    ws["C6"].font = Font(bold=True)

    # Totais por posto (opcional no Excel)
    ws["A7"] = "Cidade"
    ws["B7"] = f"TOTAL: {int(totais_posto.get('CIDADE', 0) or 0)}L"
    ws["A8"] = "Estrada"
    ws["B8"] = f"TOTAL: {int(totais_posto.get('ESTRADA', 0) or 0)}L"
    ws["A9"] = "Pontão"
    ws["B9"] = f"TOTAL: {int(totais_posto.get('PONTAO', 0) or 0)}L"

    ws["A7"].font = Font(bold=True)
    ws["A8"].font = Font(bold=True)
    ws["A9"].font = Font(bold=True)

    headers = ["Nome", "Categoria", "Equipe", "Cota", "Gasolina", "Diesel", "Total", "Saldo"]
    start_row = 11

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, d in enumerate(dados, start=start_row + 1):
        ws.cell(i, 1, d[1])
        ws.cell(i, 2, d[2])
        ws.cell(i, 3, d[3])
        ws.cell(i, 4, f"{int(d[4])}L")
        ws.cell(i, 5, f"{int(d[5])}L")
        ws.cell(i, 6, f"{int(d[6])}L")
        ws.cell(i, 7, f"{int(d[7])}L")
        ws.cell(i, 8, f"{int(d[8])}L")

    widths = [22, 18, 22, 10, 12, 12, 12, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"resumo_{mes}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/resumo/pdf")
@login_required
def resumo_pdf():
    mes = request.args.get("mes") or datetime.now().strftime("%Y-%m")
    nome_f = (request.args.get("nome") or "").strip()
    categoria_f = normalizar(request.args.get("categoria") or "")
    equipe_f = normalizar(request.args.get("equipe") or "")

    # buscar_resumo agora retorna 3 coisas:
    # dados, totais_combustivel, totais_posto
    dados, totais_combustivel, totais_posto = buscar_resumo(
        mes, nome_f, categoria_f, equipe_f, somente_com_movimento=False
    )

    gasolina_total = int(totais_combustivel.get("gasolina", 0) or 0)
    diesel_total = int(totais_combustivel.get("diesel", 0) or 0)

    # =========================================================
    # Totais por POSTO separando GASOLINA e DIESEL
    # =========================================================
    postos = {
        "gasolina": {"CIDADE": 0.0, "ESTRADA": 0.0, "PONTAO": 0.0},
        "diesel":   {"CIDADE": 0.0, "ESTRADA": 0.0, "PONTAO": 0.0},
    }

    sql = """
        SELECT
            IFNULL(e.posto,'CIDADE') as posto,
            LOWER(IFNULL(e.combustivel,'')) as combustivel,
            IFNULL(SUM(e.quantidade), 0) as total
        FROM entregas e
        JOIN cadastrados c ON e.cadastrado_id = c.id
        WHERE strftime('%Y-%m', e.data) = ?
    """
    params = [mes]

    if nome_f:
        sql += " AND c.nome LIKE ?"
        params.append(f"%{nome_f}%")
    if categoria_f:
        sql += " AND c.categoria = ?"
        params.append(categoria_f)
    if equipe_f:
        sql += " AND c.equipe = ?"
        params.append(equipe_f)

    sql += " GROUP BY IFNULL(e.posto,'CIDADE'), LOWER(IFNULL(e.combustivel,''))"

    with sqlite3.connect(DB_FILE) as conn:
        cur = conn.cursor()
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()

    for posto, comb, total in rows:
        p = (posto or "CIDADE").strip().upper()
        if p == "PONTÃO":
            p = "PONTAO"
        if p not in ("CIDADE", "ESTRADA", "PONTAO"):
            p = "CIDADE"

        c = (comb or "").strip().lower()
        if c not in ("gasolina", "diesel"):
            continue

        postos[c][p] += float(total or 0)

    gas_cidade  = int(postos["gasolina"]["CIDADE"]  or 0)
    gas_estrada = int(postos["gasolina"]["ESTRADA"] or 0)
    gas_pontao  = int(postos["gasolina"]["PONTAO"]  or 0)

    die_cidade  = int(postos["diesel"]["CIDADE"]  or 0)
    die_estrada = int(postos["diesel"]["ESTRADA"] or 0)
    die_pontao  = int(postos["diesel"]["PONTAO"]  or 0)

    log_acao("EXPORTAR_RESUMO_PDF", "resumo", None, {
        "mes": mes, "nome": nome_f, "categoria": categoria_f, "equipe": equipe_f
    })

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=25,
        rightMargin=25,
        topMargin=20,
        bottomMargin=50
    )

    def desenhar_cabecalho(canvas, doc_obj):
        w, h = doc_obj.pagesize

        # Título
        canvas.setFont("Helvetica-Bold", 18)
        canvas.drawCentredString(w / 2, h - 35, f"TOTAL MÊS {mes_pt_ano(mes).upper()}")

        # Emissão
        canvas.setFont("Helvetica", 9)
        canvas.drawString(doc_obj.leftMargin, h - 55, f"Emissão: {agora_emissao_str()}")

        # Área dos cards
        top_y = h - 70
        cards_y = top_y - 90
        gap = 18
        cards_w = (w - doc_obj.leftMargin - doc_obj.rightMargin - gap) / 2
        card_h = 78

        x1 = doc_obj.leftMargin
        x2 = doc_obj.leftMargin + cards_w + gap

        # Função helper para “card” travado e bonito
        def draw_card(x, y, title, total, c, e, p):
            # Fundo (cinza bem leve) + borda
            canvas.setFillColorRGB(0.96, 0.96, 0.96)
            canvas.setStrokeColorRGB(0.55, 0.55, 0.55)
            canvas.setLineWidth(1)
            canvas.roundRect(x, y, cards_w, card_h, 14, stroke=1, fill=1)

            # Título
            canvas.setFillColorRGB(0, 0, 0)
            canvas.setFont("Helvetica-Bold", 12)
            canvas.drawString(x + 14, y + card_h - 22, title)

            # Total grande
            canvas.setFont("Helvetica-Bold", 18)
            canvas.drawRightString(x + cards_w - 14, y + card_h - 28, f"{int(total)}L")

            # Linha divisória
            canvas.setStrokeColorRGB(0.75, 0.75, 0.75)
            canvas.setLineWidth(0.7)
            canvas.line(x + 14, y + 34, x + cards_w - 14, y + 34)

            # Detalhe por posto (travado em 3 colunas)
            canvas.setFillColorRGB(0.10, 0.10, 0.10)
            canvas.setFont("Helvetica", 10)

            col1 = x + 14
            col2 = x + (cards_w / 3) + 6
            col3 = x + (2 * cards_w / 3) + 6

            canvas.drawString(col1, y + 14, f"Cidade: {int(c)}L")
            canvas.drawString(col2, y + 14, f"Estrada: {int(e)}L")
            canvas.drawString(col3, y + 14, f"Pontão: {int(p)}L")

        # Card Gasolina
        draw_card(
            x1, cards_y,
            "GASOLINA (TOTAL)",
            gasolina_total,
            gas_cidade, gas_estrada, gas_pontao
        )

        # Card Diesel
        draw_card(
            x2, cards_y,
            "DIESEL (TOTAL)",
            diesel_total,
            die_cidade, die_estrada, die_pontao
        )

        # Filtros abaixo dos cards
        canvas.setFont("Helvetica", 9)
        filtros_txt = f"Filtros: nome='{nome_f or 'TODOS'}' | categoria='{categoria_f or 'TODOS'}' | equipe='{equipe_f or 'TODOS'}'"
        canvas.drawString(doc_obj.leftMargin, cards_y - 16, filtros_txt)

        desenhar_assinatura_pdf(canvas, doc_obj)

    # Tabela (desce um pouco por causa dos cards)
    story = [Spacer(1, 165)]
    table_data = [["Nome", "Categoria", "Equipe", "Cota", "Gasolina", "Diesel", "Total", "Saldo"]]
    for d in dados:
        table_data.append([
            d[1], d[2], d[3],
            f"{int(d[4])}L",
            f"{int(d[5])}L",
            f"{int(d[6])}L",
            f"{int(d[7])}L",
            f"{int(d[8])}L",
        ])

    col_widths = [150, 130, 140, 70, 80, 80, 80, 80]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.7, colors.grey),
        ("ALIGN", (3, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
    ]))
    story.append(table)

    doc.build(story, onFirstPage=desenhar_cabecalho)
    buffer.seek(0)

    filename = f"resumo_{mes}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    app.run(debug=True) 
